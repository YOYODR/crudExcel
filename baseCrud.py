from datetime import datetime
from openpyxl import load_workbook
rut=r'baseCrud.xlsx'

def leer(ruta:str, extraer:str):
  archivoExcel=load_workbook(ruta)

  hojaDatos=archivoExcel['Datos del crud']
  hojaDatos=hojaDatos['A2':'F'+str(hojaDatos.max_row)]

  info={}

  for i in hojaDatos:
    if isinstance(i[0].value,int):
      info.setdefault(i[0].value,{'titulo':i[1].value, 'descripcion':i[2].value,'estado':i[3].value,'fecha inicio':i[4].value,'fecha finalizacion':i[5].value})

  if not(extraer=='todo'):
    info=filtrar(info,extraer)
  for i in info:
    print('********** Tarea ***********')
    print('id:'+str(i)+'\n'+'titulo: '+str(info[i]['titulo'])+'\n'+'descripcion: '+str(info[i]['descripcion'])+'\n'+'estado: '+str(info[i]['estado'])+'\n'+'fecha de creacion: '+str(info[i]['fecha inicio'])+'\n'+'fecha finalizacion: '+str(info[i]['fecha finalizacion']))
    print()
  return info

def filtrar(info:dict, filtro:str):
  aux={}
  for i in info:
    if info[i]['estado']==filtro:
      aux.setdefault(i,info[i])
  return aux



def actualizar(ruta:str,identificador:int,datosActualizados:dict):
  archivoExcel=load_workbook(ruta)

  hojaDatos=archivoExcel['Datos del crud']
  hojaDatos=hojaDatos['A2':'F'+str(hojaDatos.max_row)]
  hoja=archivoExcel.active

  titulo=2
  descripcion=3
  estado=4
  fechaInicio=5
  fechaFinalizado=6
  encontro=False
  for i in hojaDatos:
    if i[0].value==identificador:
      fila=i[0].row
      encontro=True
      for d in datosActualizados:
        if d=='titulo' and not(datosActualizados[d]==''):
          hoja.cell(row=fila,column=titulo).value=datosActualizados[d]
        elif d=='descripcion' and not(datosActualizados[d]==''):
          hoja.cell(row=fila,column=descripcion).value=datosActualizados[d]
        elif d=='estado' and not(datosActualizados[d]==''):
          hoja.cell(row=fila,column=estado).value=datosActualizados[d]
        elif d=='fecha inicio' and not(datosActualizados[d]==''):
          hoja.cell(row=fila,column=fechaInicio).value=datosActualizados[d]
        elif d=='fecha finalizado' and not(datosActualizados[d]==''):
          hoja.cell(row=fila,column=fechaFinalizado).value=datosActualizados[d]
  archivoExcel.save(ruta)
  if encontro == False:
    print('Error: no existe una tarea con ese id')
    print()
  return

def agregar(ruta:int, datos:dict):
  archivoExcel=load_workbook(ruta)

  hojaDatos=archivoExcel['Datos del crud']
  hojaDatos=hojaDatos['A2':'F'+str(hojaDatos.max_row+1)]
  hoja=archivoExcel.active

  titulo=2
  descripcion=3
  estado=4
  fechaInicio=5
  fechaFinalizado=6
  for i in hojaDatos:
    if not(isinstance(i[0].value,int)):
      identificador=i[0].row
      hoja.cell(row=identificador,column=1).value=identificador-1
      hoja.cell(row=identificador,column=titulo).value=datos['titulo']
      hoja.cell(row=identificador,column=descripcion).value=datos['descripcion']
      hoja.cell(row=identificador,column=estado).value=datos['estado']
      hoja.cell(row=identificador,column=fechaInicio).value=datos['fecha inicio']
      hoja.cell(row=identificador,column=fechaFinalizado).value=datos['fecha finalizacion']
      break
  archivoExcel.save(ruta)
  return

def borrar(ruta,identificador):
  archivoExcel=load_workbook(ruta)

  hojaDatos=archivoExcel['Datos del crud']
  hojaDatos=hojaDatos['A2':'F'+str(hojaDatos.max_row)]
  hoja=archivoExcel.active

  titulo=2
  descripcion=3
  estado=4
  fechaInicio=5
  fechaFinalizado=6
  encontro=False
  for i in hojaDatos:
    if i[0].value==identificador:
      fila=i[0].row
      encontro=True
      hoja.cell(row=fila,column=1).value=''
      hoja.cell(row=fila,column=titulo).value=''
      hoja.cell(row=fila,column=descripcion).value=''
      hoja.cell(row=fila,column=estado).value=''
      hoja.cell(row=fila,column=fechaInicio).value=''
      hoja.cell(row=fila,column=fechaFinalizado).value=''
  archivoExcel.save(ruta)
  if encontro == False:
    print('Error: no existe una tarea con ese id')
    print()
  return

datosActualizados={'titulo':'','descripcion':'','estado':'','fecha inicio':'','fecha finalizacion':''}

